import openpyxl
import glob
import os

files = glob.glob(os.path.join(r'g:\mytecherp\MytechERP', 'Fire Fighting*xlsx'))
fpath = files[0]
wb = openpyxl.load_workbook(fpath, data_only=True)
ws = wb[wb.sheetnames[0]]

# Print rows 1-10 for all columns (especially the header area with column names)
print("=== ROWS 1-10 FULL DATA (all columns up to AU) ===")
for r in range(1, 11):
    print(f"\n--- Row {r} ---")
    for c in range(1, 48):  # Up to column AU (47)
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  {col_letter}{r} = {cell.value}")

# Get Column headers - rows 8 and 9 which are likely the column headers
print("\n=== HEADER ROWS 8-9 ===")
for r in [8, 9]:
    for c in range(1, 48):
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  {col_letter}{r} = {repr(cell.value)}")

# Merged cells
print("\n=== MERGED CELLS ===")
for mc in ws.merged_cells.ranges:
    print(f"  {mc}")

# Also check rows 31-37 for totals area
print("\n=== ROWS 31-37 (TOTALS/TAXES) ===")
for r in range(31, 38):
    print(f"\n--- Row {r} ---")
    for c in range(1, 48):
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  {col_letter}{r} = {cell.value}")

# Also read formulas for rows 31-37
wb2 = openpyxl.load_workbook(fpath, data_only=False)
ws2 = wb2[wb2.sheetnames[0]]
print("\n=== ROWS 31-37 FORMULAS ===")
for r in range(31, 38):
    print(f"\n--- Row {r} ---")
    for c in range(1, 48):
        cell = ws2.cell(row=r, column=c)
        if cell.value is not None and str(cell.value).startswith('='):
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  {col_letter}{r}: {cell.value}")

# Check what columns F/H are (imported/local)
print("\n=== Columns E-I for rows 5-37 ===")
for r in range(5, 37):
    row_data = {}
    for c in range(5, 10):  # E through I
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_data[f"{col_letter}{r}"] = cell.value
    if row_data:
        print(f"  Row {r}: {row_data}")
