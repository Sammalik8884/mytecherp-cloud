import openpyxl
import os

path = r"g:\mytecherp\MytechERP"
for f in os.listdir(path):
    if "Quotation" in f and f.endswith(".xlsx"):
        full = os.path.join(path, f)
        wb = openpyxl.load_workbook(full)
        ws = wb.active
        
        print("=== HEADER ROWS 1-10 with ALL 26 columns (formulas) ===")
        for r in range(1, 11):
            for c in range(1, 27):
                v = ws.cell(r, c).value
                if v:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    print(f"  {col_letter}{r}: {str(v)[:120]}")
            print()

        # Read data-only version for row 10 values
        wb2 = openpyxl.load_workbook(full, data_only=True)
        ws2 = wb2.active
        print("\n=== ROW 10 VALUES (data_only) ===")
        for c in range(1, 27):
            v = ws2.cell(10, c).value
            if v:
                col_letter = openpyxl.utils.get_column_letter(c)
                print(f"  {col_letter}10: {v}")
        
        print("\n=== ROW 16 as example (data_only) - all cols ===")
        for c in range(1, 27):
            v = ws2.cell(16, c).value
            if v:
                col_letter = openpyxl.utils.get_column_letter(c)
                print(f"  {col_letter}16: {v}")
        break
